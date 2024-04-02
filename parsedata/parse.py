import sys
import os
import shutil
from openpyxl import load_workbook
import json

from .helpers import jumlahkanList
from .helpers import getDetailGuru

filePath = sys.argv[2] if len(sys.argv) >= 2 else "data jadwal guru.xlsx"
wb = load_workbook(filePath)

# JSON result path
data_dir = "docs"
kelas_result_dir = os.path.join(data_dir, "kelas")
guru_result_dir = os.path.join(data_dir, "guru")
lokasi_guru_result_dir = os.path.join(data_dir, "lokasi_guru")

if os.path.isdir(kelas_result_dir):
    shutil.rmtree(kelas_result_dir)
if os.path.isdir(guru_result_dir):
    shutil.rmtree(guru_result_dir)
if os.path.isdir(lokasi_guru_result_dir):
    shutil.rmtree(lokasi_guru_result_dir)
os.makedirs(kelas_result_dir, exist_ok=True)
os.makedirs(guru_result_dir, exist_ok=True)
os.makedirs(lokasi_guru_result_dir, exist_ok=True)

# Worksheet
jadwalWs = wb["master"]
guruWs = wb["dftr nama gr"]

## Jumlah jam pelajaran / hari
jam_pelajaran = [8, 10, 10, 10, 9]
jumlah_istirahat = [1, 2, 2, 2, 2]
jumlah_jam_pelajaran = [jam + jumlah_istirahat[idx] for idx, jam in enumerate(jam_pelajaran)]

## Rentang jadwal kelas 10
# rentang_kelas_x = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
rentang_kelas_x = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

## Rentang jadwal kelas 11
rentang_kelas_xi = [kelas_x + (len(rentang_kelas_x) * 1) for kelas_x in rentang_kelas_x]

## Rentang jadwal kelas 12
rentang_kelas_xii = [kelas_x + (len(rentang_kelas_x) * 2) for kelas_x in rentang_kelas_x]
rentang_kelas_xii_mipa = rentang_kelas_xii[:8]
rentang_kelas_xii_ips = rentang_kelas_xii[8:]

# Main function: Parse xlsx
def parseJadwal(awalan_kelas: str, rentang_kelas: list):
    list_kelas = []

    for index, rentang_kolom in enumerate(rentang_kelas, start=1):
        kelas = f"{awalan_kelas}-{index}"
        list_kelas.append(kelas)

        kelas_jadwal = []
        for hari in range(1, 5+1):
            kelas_jadwal_hari = []

            lompatan_hari = jumlahkanList(jumlah_jam_pelajaran[:hari-1])
            if hari == 1:
                lompatan_hari = 0

            min_row = 11 + lompatan_hari
            if hari > 1:
                min_row += hari - 1
            max_row = min_row + jumlah_jam_pelajaran[hari-1] - 1

            row_idx = 1
            for row in jadwalWs.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=rentang_kolom,
                    max_col=rentang_kolom
                ):
                if not row[0].value:
                    continue

                jadwal = getDetailGuru(row[0].value, guruWs)
                jadwal["kelas"] = kelas
                jadwal["hari"] = hari
                jadwal["jam"] = row_idx

                # Tambah lokasi guru ke JSON
                lokasi_guru_result = os.path.join(lokasi_guru_result_dir, f"{jadwal['index']}.json")
                if os.path.isfile(lokasi_guru_result):
                    with open(lokasi_guru_result, "r") as file:
                        lokasi_guru_data = json.load(file)
                else:
                    lokasi_guru_data = []
                lokasi_guru_data.append(jadwal)
                with open(lokasi_guru_result, "w") as file:
                    json.dump(lokasi_guru_data, file, indent=4)

                kelas_jadwal_hari.append(jadwal)
                row_idx += 1
            kelas_jadwal.append(kelas_jadwal_hari)

        kelas_result = os.path.join(kelas_result_dir, f"{kelas}.json")
        with open(kelas_result, "w") as file:
            json.dump(kelas_jadwal, file, indent=4)

    # Daftar untuk web
    list_kelas_result = os.path.join(kelas_result_dir, "info.json")
    if os.path.isfile(list_kelas_result):
        with open(list_kelas_result, "r") as file:
            list_kelas_data = json.load(file)
    else:
        list_kelas_data = []
    list_kelas_data.append({
        "jenjang": awalan_kelas,
        "kelas": list_kelas
    })
    with open(list_kelas_result, "w") as file:
        json.dump(list_kelas_data, file, indent=4)

# Main function: Get all guru
def getAllGuru():
    guru_list = []

    for row in guruWs.iter_rows(min_row=3, max_row=guruWs.max_row):
        index_guru = row[0].value
        nama_guru = row[1].value
        keterangan_guru = row[3].value

        guru_info = {
            "index": index_guru,
            "nama": nama_guru,
            "keterangan": keterangan_guru
        }
        guru_list.append(guru_info)

    guru_result = os.path.join(guru_result_dir, "all.json")
    with open(guru_result, "w") as file:
        json.dump(guru_list, file, indent=4)

#####

def startParse():
    # Get all guru
    getAllGuru()

    # Parse kelas 10
    parseJadwal("X", rentang_kelas_x)

    # Parse kelas 11
    parseJadwal("XI", rentang_kelas_xi)

    # Parse kelas 12 - MIPA
    parseJadwal("XII-MIPA", rentang_kelas_xii_mipa)

    # Parse kelas 13 - IPS
    parseJadwal("XII-IPS", rentang_kelas_xii_ips)
