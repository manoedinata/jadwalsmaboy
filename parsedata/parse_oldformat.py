import sys
import os
import shutil
from openpyxl import load_workbook
import json

from .helpers_oldformat import getDetailGuru

filePath = sys.argv[1] if len(sys.argv) >= 2 else "Data Pelajaran.xlsx"
wb = load_workbook(filePath)

# JSON result path
data_dir = "docs"
kelas_result_dir = os.path.join(data_dir, "kelas")
guru_result_dir = os.path.join(data_dir, "guru")
lokasi_guru_result_dir = os.path.join(data_dir, "lokasi_guru")
ujian_result_dir = os.path.join(data_dir, "ujian")

if os.path.isdir(kelas_result_dir):
    shutil.rmtree(kelas_result_dir)
if os.path.isdir(guru_result_dir):
    shutil.rmtree(guru_result_dir)
if os.path.isdir(lokasi_guru_result_dir):
    shutil.rmtree(lokasi_guru_result_dir)
if os.path.isdir(ujian_result_dir):
    shutil.rmtree(ujian_result_dir)
os.makedirs(kelas_result_dir, exist_ok=True)
os.makedirs(guru_result_dir, exist_ok=True)
os.makedirs(lokasi_guru_result_dir, exist_ok=True)
os.makedirs(ujian_result_dir, exist_ok=True)

# Worksheet
LIST_KELAS = [("X", 1), ("XI", 2), ("XII", 3)]
GURU_WORKSHEET = wb["Data Guru"]
GURU_WORKSHEET_ROWS = list(GURU_WORKSHEET.rows)
UJIAN_WORKSHEET = wb["Ujian"]

## Jumlah jam pelajaran / hari
HARI = [8, 10, 10, 10, 9]
MAX_JAMPELAJARAN = 10

# Lokasi kolom ujian
kolom_ujian_x = 2 # C
kolom_ujian_xi_tipe1 = 3 # D
kolom_ujian_xi_tipe2 = 4 # E
kolom_ujian_xi_tipe3 = 5 # F
kolom_ujian_xi_tipe4 = 6 # G
kolom_ujian_xi_tipe5 = 7 # H
kolom_ujian_xii_ipa = 8 # I
kolom_ujian_xii_ips = 9 # J

# Main function: Parse xlsx
def parseJadwal(anggotaKelas):
    kelas = wb[anggotaKelas[0]]
    kelasCol = list(kelas.columns)

    listKelas = []
    for k in kelasCol:
        namaKelas = k[0].value # Nama kelas
        if not namaKelas: continue
        listKelas.append(namaKelas)
        print(namaKelas)

        jadwalPure = k[1:]
        penandaHari = 1
        i = 1

        kelas_jadwal = []
        kelas_jadwal_hari = []
        for j in jadwalPure:
            if j.value != None:
                # Reminder:
                # j.value = Indeks guru
                # i = Jam pelajaran
                jadwal = getDetailGuru(j.value, GURU_WORKSHEET)
                jadwal["kelas"] = namaKelas
                jadwal["hari"] = penandaHari
                jadwal["jam"] = i

                # Tambah lokasi guru ke JSON
                lokasi_guru_result = os.path.join(
                    lokasi_guru_result_dir,
                    f"{j.value}.json"
                )
                lokasi_guru_data = []
                if os.path.isfile(lokasi_guru_result):
                    with open(lokasi_guru_result, "r") as file:
                        lokasi_guru_data = json.load(file)
                lokasi_guru_data.append(jadwal)
                with open(lokasi_guru_result, "w") as file:
                    json.dump(lokasi_guru_data, file, indent=4)

                # Tambahkan jadwal ke kelas_jadwal_hari
                kelas_jadwal_hari.append(jadwal)

                i += 1
            else:
                penandaHari += 1
                i = 1
                kelas_jadwal.append(kelas_jadwal_hari)
                kelas_jadwal_hari = []
                continue

            # Reset penandaHari di setiap for loop
            # penandaHari = 1

        # Dengan metode penambahan data array di atas,
        # data terakhir kemungkinan besar belum ditambahkan ke array.
        # Tambahkan secara manual.
        kelas_jadwal.append(kelas_jadwal_hari)

        kelas_result = os.path.join(kelas_result_dir, f"{namaKelas}.json")
        with open(kelas_result, "w") as file:
            json.dump(kelas_jadwal, file, indent=4)

    # Daftar untuk web
    list_kelas_result = os.path.join(kelas_result_dir, "info.json")
    if os.path.isfile(list_kelas_result):
        with open(list_kelas_result, "r") as file:
            list_kelas_data = json.load(file)
    else:
        list_kelas_data = []
    list_kelas_data.append({
        "jenjang": anggotaKelas[0],
        "kelas": listKelas
    })
    with open(list_kelas_result, "w") as file:
        json.dump(list_kelas_data, file, indent=4)

# Main function: Get all guru
def getAllGuru():
    guru_list = []
    for row in GURU_WORKSHEET_ROWS[1:]:
        indeks = row[0].value
        nama = row[1].value.strip()
        mapel = row[2].value.strip()

        # Skip empty row
        if not nama:
            continue

        guru_info = {
            "index": indeks,
            "nama": nama,
            "keterangan": mapel
        }
        guru_list.append(guru_info)

    guru_result = os.path.join(guru_result_dir, "all.json")
    with open(guru_result, "w") as file:
        json.dump(guru_list, file, indent=4)

    return guru_list

def parseUjian(nama_kelas: str, kolom_kelas: int):
    total_jadwal = []
    for row in UJIAN_WORKSHEET.iter_rows(
            min_row=2
        ):

        if not row[kolom_kelas].value:
            continue

        if row[0].value:
            total_jadwal.append({
                "tanggal": row[0].value,
                "jadwal": [[row[1].value.split("-"), row[kolom_kelas].value.strip()]]
            })
        else:
            total_jadwal[-1]["jadwal"].append([row[1].value.split("-"), row[kolom_kelas].value.strip()])

    ujian_result = os.path.join(ujian_result_dir, f"{nama_kelas}.json")
    with open(ujian_result, "w") as file:
        json.dump(total_jadwal, file, indent=4)
#####

def startParse():
    # Get all guru
    getAllGuru()

    # Parse jadwal kelas
    for kelas in LIST_KELAS:
        parseJadwal(kelas)

    # Parse Ujian - kelas 10
    parseUjian("X", kolom_ujian_x)

    # Parse Ujian - kelas 11 (gathel akeh nemen)
    parseUjian("XI-1 – XI-4", kolom_ujian_xi_tipe1)
    parseUjian("XI-5 – XI-6", kolom_ujian_xi_tipe2)
    parseUjian("XI-7 – XI-8", kolom_ujian_xi_tipe3)
    parseUjian("XI-9 – XI-10", kolom_ujian_xi_tipe4)
    parseUjian("XI-11 – XI-12", kolom_ujian_xi_tipe5)

    # Parse Ujian - kelas 12
    parseUjian("XII-1 – XII-8", kolom_ujian_xii_ipa)
    parseUjian("XII-9 – XII-12", kolom_ujian_xii_ips)
